import srsly
from pathlib import Path
import typer
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

app = typer.Typer()

def strip_quotes(text):
    return text.strip().strip('"')

def clean_entities(entities, label):
    return '; '.join(ent['text'].replace('\n', ' ') for ent in entities if ent['label'] == label)

def export_to_excel(json_file: Path, output_folder: Path):
    if not output_folder.exists():
        output_folder.mkdir(parents=True, exist_ok=True)

    data = list(srsly.read_jsonl(json_file))
    rows = []
    for item in data:
        row = {
            "image": item.get('image', 'No image available'),
            "original_text": strip_quotes(item.get('text', 'No original text available')),
            "cleaned_text": strip_quotes(item.get('cleaned_text', 'No cleaned text available')),
            "english_translation": strip_quotes(item.get('english_translation', 'No translation available')),
            "summary": strip_quotes(item.get('summary', 'No summary available')),
            "persons": clean_entities(item.get('entities', []), 'PER'),
            "locations": clean_entities(item.get('entities', []), 'LOC'),
            "organizations": clean_entities(item.get('entities', []), 'ORG')
        }
        rows.append(row)

    df = pd.DataFrame(rows)
    excel_file = output_folder / "exported_data.xlsx"
    df.to_excel(excel_file, index=False)

    # Apply formatting to the Excel file
    wb = load_workbook(excel_file)
    ws = wb.active

    # Set specific column widths
    column_widths = {
        "A": 200 / 7,  # Image
        "B": 400 / 7,  # Original Text
        "C": 400 / 7,  # Cleaned Text
        "D": 400 / 7,  # English Translation
        "E": 200 / 7,  # Summary
        "F": 200 / 7,  # Persons
        "G": 200 / 7,  # Locations
        "H": 200 / 7   # Organizations
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Apply word wrap, auto height, and top left alignment to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            ws.row_dimensions[cell.row].height = None

    # Add filter to the header row
    ws.auto_filter.ref = ws.dimensions

    wb.save(excel_file)
    print(f"Excel file saved as {excel_file}")

@app.command()
def main(
    json_file: Path = typer.Argument(..., help="Path to the JSONL file"),
    output_folder: Path = typer.Argument(..., help="Path to the output folder for Excel file")
):
    export_to_excel(json_file, output_folder)

if __name__ == "__main__":
    typer.run(main)